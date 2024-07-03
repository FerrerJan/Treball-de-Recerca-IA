from openpyxl import Workbook
from openpyxl import load_workbook
import random

with open("/media/teo/1T Disc Dur/TRABAJO DE RECERCA/Treball-de-Recerca-IA/TR/DATOS/datos.txt") as f:
    contents = f.readlines()
    matrix = []
    for i in range(len(contents)):
        matrix.append([eval(i) for i in (contents[i].replace("\n", "").replace(",", "").split())])
#acceso al archivo .txt i  lectura de datos para una posterior transformación  en una matriz
        
lista_promedio = []
for i in range(len(matrix[0])):
    n = 0
    for f in range(len(matrix)):
        n += matrix[f][i]
    lista_promedio.append(round((n/len(matrix)), 2))
#creación de una lista que contiene un promedio de todos los valores

wb = load_workbook(filename = 'IA.xlsx')
print(wb.sheetnames)

columnas_excel = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
#workbook = Workbook()
#sheet = workbook.active
 
#definicion de variables

nw = wb["Sheet"]
if nw["A1"].value != None:
    name = "new" + str(random.randint(1,10))
    new = wb.create_sheet(name)
    new = wb[name]
    for i in range(len(lista_promedio)):
        new[columnas_excel[0] + str(i+1)] = i+1
        new[columnas_excel[1] + str(i+1)] = lista_promedio[i]
#creación de una tabla en un excel con los valores medios de fitness de cada generacion
else:
    for i in range(len(lista_promedio)):
        nw[columnas_excel[0] + str(i+1)] = i+1
        nw[columnas_excel[1] + str(i+1)] = lista_promedio[i]
       
#creación de una nueva hoja en caso de  que la hoja activa no esté vacía

wb.save(filename="IA.xlsx")
#guardado del archivo

"""
    from openpyxl import load_workbook
    wb = load_workbook(filename = 'empty_book.xlsx')
    sheet_ranges = wb['range names']
    print(sheet_ranges['D18'].value)
    3
"""